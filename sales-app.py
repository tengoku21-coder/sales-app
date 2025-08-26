
# -*- coding: utf-8 -*-
"""
Streamlit 영업 보고서 관리 (웹버전)
- 기본 영업자 '김범준' (변경 가능)
- 날짜 기본값: 오늘
- 연락처 자동 하이픈(02/010 규칙)
- 충전기 모델/부대공사 수량 + 기타모델 추가/삭제
- 미리보기 자동 생성
- 엑셀(.xlsx) 저장/수정/삭제/다운로드
- 저장 목록 표시는 ID 숨김(내부적으로만 사용)
"""
import io
import json
import os
from datetime import date, datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook

# -------------------------------
# 설정
# -------------------------------
APP_TITLE = "영업 보고서 관리 (웹)"
DEFAULT_EXCEL_PATH = os.environ.get("SR_EXCEL_PATH", "영업보고서.xlsx")

BUILTIN_MODELS = [
    "2100A","1100A","3050A","3050B","3050C","2007CP","2007A","2007C","1030A"
]
ACCESSORY_ITEMS = [
    "i형볼라드","u형 볼라드","기초패드(완속)","기초패드(급속)",
    "바닥면도색","캐노피(완속)","캐노피(급속)"
]

EXCEL_COLUMNS = [
    "ID","날짜","영업자","현장명","담당자","연락처",
    "진행상태","불가사유","비고",
] + [f"모델_{m}" for m in BUILTIN_MODELS] \
  + [f"자재_{a}" for a in ACCESSORY_ITEMS] \
  + ["기타(JSON)","충전기요약"]


# -------------------------------
# 유틸
# -------------------------------
def ensure_workbook(path: str = DEFAULT_EXCEL_PATH):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"
        for i, col in enumerate(EXCEL_COLUMNS, start=1):
            ws.cell(row=1, column=i, value=col)
        wb.save(path)


def retrofit_missing_ids(path: str = DEFAULT_EXCEL_PATH):
    """ID 없는 기존 행에 ID 부여(삭제/수정 안정화)."""
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx_map = {h: i + 1 for i, h in enumerate(headers) if h}
    id_col = idx_map.get("ID")
    if not id_col:
        return
    changed = False
    for r in range(2, ws.max_row + 1):
        if not ws.cell(row=r, column=id_col).value:
            new_id = f"SR-{datetime.now():%Y%m%d%H%M%S}-{str(r).zfill(4)}"
            ws.cell(row=r, column=id_col, value=new_id)
            changed = True
    if changed:
        wb.save(path)


def load_all(path: str = DEFAULT_EXCEL_PATH):
    ensure_workbook(path)
    retrofit_missing_ids(path)
    wb = load_workbook(path)
    ws = wb.active
    data = ws.values
    header = next(data)
    rows = list(data)
    df = pd.DataFrame(rows, columns=header)
    df = df.dropna(how="all")
    return df


def append_record(record: dict, path: str = DEFAULT_EXCEL_PATH):
    wb = load_workbook(path)
    ws = wb.active
    next_row = ws.max_row + 1
    for i, col in enumerate(EXCEL_COLUMNS, start=1):
        ws.cell(row=next_row, column=i, value=record.get(col))
    wb.save(path)


def update_record(record: dict, path: str = DEFAULT_EXCEL_PATH) -> bool:
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx_map = {h: i + 1 for i, h in enumerate(headers) if h}
    id_col = idx_map.get("ID")
    if not id_col:
        return False
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=id_col).value) == str(record.get("ID")):
            for i, col in enumerate(EXCEL_COLUMNS, start=1):
                ws.cell(row=r, column=i, value=record.get(col))
            wb.save(path)
            return True
    return False


def delete_records(ids, path: str = DEFAULT_EXCEL_PATH) -> int:
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx_map = {h: i + 1 for i, h in enumerate(headers) if h}
    id_col = idx_map.get("ID")
    if not id_col:
        return 0
    ids = set(map(str, ids))
    deleted = 0
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(row=r, column=id_col).value) in ids:
            ws.delete_rows(r, 1)
            deleted += 1
    wb.save(path)
    return deleted


def format_phone(s: str) -> str:
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits.startswith("02"):
        if len(digits) == 10:
            return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
        elif len(digits) == 9:
            return f"{digits[:2]}-{digits[2:5]}-{digits[5:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:]}"
    if len(digits) > 7:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) > 3:
        return f"{digits[:3]}-{digits[3:]}"
    return digits


def build_summary(model_counts: dict, accessory_counts: dict, others: list) -> str:
    parts = []
    for k, v in model_counts.items():
        if v > 0:
            parts.append(f"{k} x {v}")
    for k, v in accessory_counts.items():
        if v > 0:
            parts.append(f"{k} x {v}")
    for name, qty in others:
        parts.append(f"{name} x {qty}")
    if not parts:
        return "— (수량 없음)"
    total = sum(int(p.split(' x ')[1]) for p in parts if ' x ' in p)
    return ", ".join(parts) + f"  | 총 {total}개"


def get_download_xlsx_bytes(path: str = DEFAULT_EXCEL_PATH) -> bytes:
    wb = load_workbook(path)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -------------------------------
# Streamlit UI
# -------------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

# 세션 상태 초기화
if "others" not in st.session_state:
    st.session_state.others = []
if "edit_id" not in st.session_state:
    st.session_state.edit_id = None

for m in BUILTIN_MODELS:
    st.session_state.setdefault(f"m_{m}", 0)
for a in ACCESSORY_ITEMS:
    st.session_state.setdefault(f"a_{a}", 0)

# 좌우 레이아웃
left, right = st.columns([2, 1])

with left:
    st.subheader("기본 정보")
    c1, c2, c3, c4 = st.columns([1,1,1,1])
    d_date = c1.date_input("날짜", value=st.session_state.get("date", date.today()), key="date")
    sales = c2.text_input("영업자", value=st.session_state.get("sales", "김범준"), key="sales")
    site = c3.text_input("현장명", value=st.session_state.get("site", ""), key="site")
    person = c4.text_input("담당자", value=st.session_state.get("person", ""), key="person")

    phone_raw = st.text_input("연락처", value=st.session_state.get("phone", ""), key="phone")
    phone = format_phone(phone_raw)
    if phone != phone_raw:
        st.session_state.phone = phone  # 즉시 보정

    status = st.radio("진행 상태", ["진행중","완료","불가"], horizontal=True, index={"진행중":0,"완료":1,"불가":2}[st.session_state.get("status","진행중")], key="status")
    reason = st.text_input("불가 사유", value=st.session_state.get("reason",""), key="reason", disabled=(status!="불가"))

    memo = st.text_area("비고", height=100, value=st.session_state.get("memo",""), key="memo")

    st.markdown("---")
    st.subheader("내용: 충전기/부대공사 수량 입력")

    cols = st.columns(3)
    model_counts = {}
    for i, m in enumerate(BUILTIN_MODELS):
        with cols[i % 3]:
            model_counts[m] = st.number_input(m, min_value=0, max_value=999, value=int(st.session_state.get(f"m_{m}",0)), step=1, key=f"m_{m}")

    st.divider()

    cols2 = st.columns(3)
    accessory_counts = {}
    for i, a in enumerate(ACCESSORY_ITEMS):
        with cols2[i % 3]:
            accessory_counts[a] = st.number_input(a, min_value=0, max_value=999, value=int(st.session_state.get(f"a_{a}",0)), step=1, key=f"a_{a}")

    st.markdown("##### 기타 모델 추가")
    oc1, oc2, oc3 = st.columns([2,1,1])
    with oc1:
        other_name = st.text_input("기타 모델명", key="other_name")
    with oc2:
        other_qty = st.number_input("수량", min_value=0, max_value=999, value=0, key="other_qty")
    with oc3:
        if st.button("추가"):
            if other_name and other_qty > 0:
                st.session_state.others.append((other_name, int(other_qty)))
                st.session_state.other_name = ""
                st.session_state.other_qty = 0
            else:
                st.warning("기타 모델명과 수량을 확인하세요.")

    if st.session_state.others:
        st.write("**기타 목록**")
        df_others = pd.DataFrame(st.session_state.others, columns=["모델","수량"])
        st.dataframe(df_others, use_container_width=True, hide_index=True)
        remove_idx = st.selectbox(
            "삭제할 항목 선택",
            options=[("선택하세요", None)] + [(f"{n} x {q}", i) for i, (n,q) in enumerate(st.session_state.others)],
            format_func=lambda x: x[0],
            index=0
        )
        cdel1, cdel2 = st.columns([1,3])
        if cdel1.button("선택 삭제"):
            if remove_idx[1] is not None:
                st.session_state.others.pop(remove_idx[1])
            else:
                st.warning("삭제할 항목을 선택하세요.")
        if cdel2.button("모두 초기화"):
            st.session_state.others.clear()

    st.markdown("---")
    col_b1, col_b2, col_b3 = st.columns([1,1,1])
    if col_b1.button("엑셀 저장(신규)"):
        summary = build_summary(model_counts, accessory_counts, st.session_state.others)
        if site.strip() == "":
            st.error("현장명을 입력하세요.")
        else:
            rec = {
                "ID": f"SR-{datetime.now():%Y%m%d%H%M%S}",
                "날짜": d_date.strftime("%Y-%m-%d"),
                "영업자": sales.strip() or "김범준",
                "현장명": site.strip(),
                "담당자": person.strip(),
                "연락처": phone,
                "진행상태": status,
                "불가사유": reason if status == "불가" else "",
                "비고": memo.strip(),
                "기타(JSON)": json.dumps(st.session_state.others, ensure_ascii=False),
                "충전기요약": summary,
            }
            for m, q in model_counts.items():
                rec[f"모델_{m}"] = int(q or 0)
            for a, q in accessory_counts.items():
                rec[f"자재_{a}"] = int(q or 0)
            ensure_workbook(DEFAULT_EXCEL_PATH)
            append_record(rec, DEFAULT_EXCEL_PATH)
            st.success("신규 저장 완료!")

    if col_b2.button("선택 수정"):
        if not st.session_state.edit_id:
            st.warning("먼저 우측 목록에서 행을 선택하고 불러오기를 누르세요.")
        else:
            summary = build_summary(model_counts, accessory_counts, st.session_state.others)
            rec = {
                "ID": st.session_state.edit_id,
                "날짜": d_date.strftime("%Y-%m-%d"),
                "영업자": sales.strip() or "김범준",
                "현장명": site.strip(),
                "담당자": person.strip(),
                "연락처": phone,
                "진행상태": status,
                "불가사유": reason if status == "불가" else "",
                "비고": memo.strip(),
                "기타(JSON)": json.dumps(st.session_state.others, ensure_ascii=False),
                "충전기요약": summary,
            }
            for m in BUILTIN_MODELS:
                rec[f"모델_{m}"] = int(st.session_state.get(f"m_{m}",0) or 0)
            for a in ACCESSORY_ITEMS:
                rec[f"자재_{a}"] = int(st.session_state.get(f"a_{a}",0) or 0)
            ok = update_record(rec, DEFAULT_EXCEL_PATH)
            if ok:
                st.success("수정 완료!")
                st.session_state.edit_id = None
            else:
                st.error("수정 실패: 해당 ID를 찾을 수 없습니다.")

    if col_b3.button("폼 초기화"):
        for key in ["sales","site","person","phone","memo","status","reason"]:
            st.session_state[key] = "" if key != "status" else "진행중"
        st.session_state.date = date.today()
        for m in BUILTIN_MODELS:
            st.session_state[f"m_{m}"] = 0
        for a in ACCESSORY_ITEMS:
            st.session_state[f"a_{a}"] = 0
        st.session_state.others.clear()
        st.session_state.edit_id = None
        st.rerun()

with right:
    st.subheader("종합 미리보기")
    model_counts = {m: int(st.session_state.get(f"m_{m}", 0)) for m in BUILTIN_MODELS}
    accessory_counts = {a: int(st.session_state.get(f"a_{a}", 0)) for a in ACCESSORY_ITEMS}
    preview_text = build_summary(model_counts, accessory_counts, st.session_state.others)
    st.text_area("충전기 종류 및 수량", value=preview_text, height=150)

    st.markdown("---")
    st.subheader("저장 목록")

    df = load_all(DEFAULT_EXCEL_PATH)
    if not df.empty:
        df_show = df.copy()
        if "ID" in df_show.columns:
            df_show = df_show.drop(columns=["ID"])
        st.dataframe(df_show, use_container_width=True, hide_index=True)

        options = [("— 행 선택 —", None)] + [
            (f"{r['날짜']} | {r['현장명']} | {r.get('담당자','')}", str(r["ID"])) for _, r in df.iterrows()
        ]
        picked = st.selectbox("행 선택 (불러오기/삭제 대상)", options=options, format_func=lambda x: x[0], index=0)
        picked_id = picked[1]

        c1, c2, c3 = st.columns([1,1,2])
        if c1.button("불러오기"):
            if not picked_id:
                st.warning("먼저 행을 선택하세요.")
            else:
                row = df[df["ID"].astype(str) == str(picked_id)].iloc[0]
                st.session_state.edit_id = str(picked_id)
                try:
                    st.session_state.date = date.fromisoformat(str(row.get("날짜","")))
                except Exception:
                    st.session_state.date = date.today()
                st.session_state.sales = str(row.get("영업자","") or "")
                st.session_state.site = str(row.get("현장명","") or "")
                st.session_state.person = str(row.get("담당자","") or "")
                st.session_state.phone = str(row.get("연락처","") or "")
                st.session_state.status = str(row.get("진행상태","진행중") or "진행중")
                st.session_state.reason = str(row.get("불가사유","") or "")
                st.session_state.memo = str(row.get("비고","") or "")
                for m in BUILTIN_MODELS:
                    st.session_state[f"m_{m}"] = int(row.get(f"모델_{m}", 0) or 0)
                for a in ACCESSORY_ITEMS:
                    st.session_state[f"a_{a}"] = int(row.get(f"자재_{a}", 0) or 0)
                try:
                    st.session_state.others = json.loads(row.get("기타(JSON)","[]")) or []
                except Exception:
                    st.session_state.others = []
                st.info("좌측 폼에 값이 반영되었습니다. 수정 후 '선택 수정'으로 저장하세요.")
                st.rerun()

        if c2.button("삭제"):
            if not picked_id:
                st.warning("삭제할 행을 선택하세요.")
            else:
                n = delete_records([picked_id], DEFAULT_EXCEL_PATH)
                if n > 0:
                    st.success("삭제 완료")
                    st.rerun()
                else:
                    st.error("삭제 실패")

    else:
        st.info("저장된 항목이 없습니다. 좌측에서 입력 후 저장하세요.")

    st.download_button(
        "엑셀 내려받기",
        data=get_download_xlsx_bytes(DEFAULT_EXCEL_PATH),
        file_name="영업보고서.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="백업용으로 다운로드하세요(스트림릿 클라우드는 파일시스템이 초기화될 수 있음)."
    )
